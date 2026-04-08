"""
Module d'export des notes annexes vers Excel.

Ce module permet d'exporter les notes calculées dans un fichier Excel formaté.
"""

import logging
from typing import Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exporteur de notes annexes vers Excel."""
    
    def __init__(self, fichier_sortie: str = None):
        """
        Initialise l'exporteur avec le nom du fichier de sortie.
        
        Args:
            fichier_sortie: Nom du fichier Excel (optionnel, généré automatiquement si None)
        """
        if fichier_sortie is None:
            date_str = datetime.now().strftime("%Y%m%d")
            fichier_sortie = f"Notes_Annexes_Calculees_{date_str}.xlsx"
        
        self.fichier_sortie = fichier_sortie
        self.workbook = Workbook()
        # Supprimer la feuille par défaut
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        logger.info(f"ExcelExporter initialisé: {fichier_sortie}")
    
    def exporter_note(self, df: pd.DataFrame, nom_onglet: str, 
                     colonnes_config: Dict = None):
        """
        Exporte une note dans un onglet Excel.
        
        Args:
            df: DataFrame de la note
            nom_onglet: Nom de l'onglet
            colonnes_config: Configuration des colonnes
        """
        # Créer l'onglet
        ws = self.workbook.create_sheet(title=nom_onglet)
        
        # Écrire les en-têtes
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Écrire les données
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Formater les montants
                if col_idx > 1:  # Colonnes de montants
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
        
        # Appliquer les bordures
        self.appliquer_formatage(ws, colonnes_config)
        
        logger.info(f"✓ Note exportée dans l'onglet '{nom_onglet}'")
    
    def exporter_toutes_notes(self, notes: Dict[str, pd.DataFrame]):
        """
        Exporte toutes les notes dans un seul fichier Excel.
        
        Args:
            notes: Dict mappant nom_note -> DataFrame
        """
        for nom_note, df in notes.items():
            self.exporter_note(df, nom_note)
        
        logger.info(f"✓ {len(notes)} notes exportées")
    
    def appliquer_formatage(self, worksheet, colonnes_config: Dict = None):
        """Applique bordures, couleurs, formats numériques."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Ajuster la largeur des colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def sauvegarder(self):
        """Sauvegarde le fichier Excel."""
        try:
            self.workbook.save(self.fichier_sortie)
            logger.info(f"✓ Fichier Excel sauvegardé: {self.fichier_sortie}")
            return True
        except Exception as e:
            logger.error(f"✗ Erreur lors de la sauvegarde: {e}")
            return False
