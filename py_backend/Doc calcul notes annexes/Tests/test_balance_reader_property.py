"""
Property-Based Tests for Balance_Reader Module

This module contains property-based tests using Hypothesis to verify
the Balance_Reader module's behavior across a wide range of inputs.

**Validates: Requirements 1.1, 1.2**

Auteur: Système de calcul automatique des notes annexes SYSCOHADA
Date: 08 Avril 2026
"""

import sys
import os
import pytest
from hypothesis import given, strategies as st, assume, settings
import pandas as pd
import openpyxl
from openpyxl import Workbook
import tempfile
import re

# Ajouter le chemin des modules au PYTHONPATH
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'Modules'))

from balance_reader import BalanceReader, BalanceNotFoundException, InvalidBalanceFormatException


# ============================================================================
# HYPOTHESIS STRATEGIES
# ============================================================================

@st.composite
def st_balance_excel_file(draw):
    """
    Génère un fichier Excel valide avec 3 onglets de balances.
    
    Cette stratégie crée un fichier Excel temporaire avec:
    - 3 onglets nommés "BALANCE N", "BALANCE N-1", "BALANCE N-2"
    - Chaque onglet contient exactement 8 colonnes
    - Les données sont cohérentes (montants numériques valides)
    
    Returns:
        str: Chemin vers le fichier Excel temporaire créé
    """
    # Générer le nombre de comptes pour chaque balance
    num_comptes_n = draw(st.integers(min_value=5, max_value=50))
    num_comptes_n1 = draw(st.integers(min_value=5, max_value=50))
    num_comptes_n2 = draw(st.integers(min_value=5, max_value=50))
    
    # Créer un fichier temporaire
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False)
    temp_file.close()
    
    # Créer le workbook
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Créer les 3 onglets
    for exercice, num_comptes in [('N', num_comptes_n), ('N-1', num_comptes_n1), ('N-2', num_comptes_n2)]:
        ws = wb.create_sheet(f"BALANCE {exercice}")
        
        # En-têtes (exactement 8 colonnes)
        headers = ['Numéro', 'Intitulé', 'Ant Débit', 'Ant Crédit', 
                   'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        ws.append(headers)
        
        # Générer les données
        for i in range(num_comptes):
            # Numéro de compte SYSCOHADA valide
            classe = draw(st.sampled_from(['1', '2', '3', '4', '5', '6', '7', '8', '9']))
            sous_classe = draw(st.integers(min_value=0, max_value=9))
            detail = draw(st.integers(min_value=0, max_value=999))
            numero = f"{classe}{sous_classe}{detail:03d}"
            
            # Intitulé
            intitule = f"Compte {numero}"
            
            # Générer les soldes d'ouverture
            ant_debit = draw(st.floats(min_value=0, max_value=1000000, allow_nan=False, allow_infinity=False))
            ant_credit = draw(st.floats(min_value=0, max_value=1000000, allow_nan=False, allow_infinity=False))
            
            # Générer les mouvements
            mvt_debit = draw(st.floats(min_value=0, max_value=500000, allow_nan=False, allow_infinity=False))
            mvt_credit = draw(st.floats(min_value=0, max_value=500000, allow_nan=False, allow_infinity=False))
            
            # Calculer les soldes de clôture cohérents
            solde_ouverture = ant_debit - ant_credit
            solde_cloture = solde_ouverture + mvt_debit - mvt_credit
            
            solde_debit = max(0, solde_cloture)
            solde_credit = max(0, -solde_cloture)
            
            # Ajouter la ligne
            ws.append([numero, intitule, ant_debit, ant_credit, 
                      mvt_debit, mvt_credit, solde_debit, solde_credit])
    
    # Sauvegarder le fichier
    wb.save(temp_file.name)
    
    return temp_file.name


# ============================================================================
# PROPERTY-BASED TESTS
# ============================================================================

@given(fichier_excel=st_balance_excel_file())
@settings(max_examples=20, deadline=30000)
def test_property_balance_loading_completeness(fichier_excel):
    """
    **Property 1: Balance Loading Completeness**
    
    **Validates: Requirements 1.1, 1.2**
    
    For any valid Excel file containing balance sheets, when the Balance_Reader 
    loads the file, all three worksheets (BALANCE N, BALANCE N-1, BALANCE N-2) 
    must be detected and loaded with exactly 8 columns each.
    
    This property verifies that:
    1. All 3 worksheets are successfully detected
    2. All 3 balances are loaded without errors
    3. Each balance has exactly 8 columns
    4. The column names are normalized correctly
    5. All balances contain at least 1 row of data
    6. All monetary columns are numeric
    7. The Numéro column is of string type
    
    Test Strategy:
    - Generate random Excel files with 3 balance sheets
    - Each sheet has 5-50 accounts with random but coherent data
    - Verify that Balance_Reader correctly loads all sheets
    - Verify column structure and data types
    """
    try:
        # Créer le lecteur
        reader = BalanceReader(fichier_excel)
        
        # Charger les balances
        balance_n, balance_n1, balance_n2 = reader.charger_balances()
        
        # Vérifier que les 3 balances sont chargées
        assert balance_n is not None, "Balance N ne doit pas être None"
        assert balance_n1 is not None, "Balance N-1 ne doit pas être None"
        assert balance_n2 is not None, "Balance N-2 ne doit pas être None"
        
        # Vérifier que chaque balance est un DataFrame
        assert isinstance(balance_n, pd.DataFrame), "Balance N doit être un DataFrame"
        assert isinstance(balance_n1, pd.DataFrame), "Balance N-1 doit être un DataFrame"
        assert isinstance(balance_n2, pd.DataFrame), "Balance N-2 doit être un DataFrame"
        
        # Vérifier que chaque balance a exactement 8 colonnes
        assert len(balance_n.columns) == 8, \
            f"Balance N doit avoir 8 colonnes, trouvé {len(balance_n.columns)}: {list(balance_n.columns)}"
        assert len(balance_n1.columns) == 8, \
            f"Balance N-1 doit avoir 8 colonnes, trouvé {len(balance_n1.columns)}: {list(balance_n1.columns)}"
        assert len(balance_n2.columns) == 8, \
            f"Balance N-2 doit avoir 8 colonnes, trouvé {len(balance_n2.columns)}: {list(balance_n2.columns)}"
        
        # Vérifier les noms de colonnes attendus
        colonnes_attendues = {
            'Numéro', 'Intitulé', 
            'Ant Débit', 'Ant Crédit',
            'Débit', 'Crédit',
            'Solde Débit', 'Solde Crédit'
        }
        
        for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
            colonnes_balance = set(balance.columns)
            assert colonnes_balance == colonnes_attendues, \
                f"Balance {nom} doit avoir les colonnes {colonnes_attendues}, trouvé {colonnes_balance}"
        
        # Vérifier que chaque balance contient au moins 1 ligne
        assert len(balance_n) > 0, "Balance N doit contenir au moins 1 ligne"
        assert len(balance_n1) > 0, "Balance N-1 doit contenir au moins 1 ligne"
        assert len(balance_n2) > 0, "Balance N-2 doit contenir au moins 1 ligne"
        
        # Vérifier que tous les montants sont numériques
        colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        
        for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
            for col in colonnes_montants:
                assert pd.api.types.is_numeric_dtype(balance[col]), \
                    f"Colonne {col} de Balance {nom} doit être numérique, trouvé {balance[col].dtype}"
                
                # Vérifier qu'il n'y a pas de NaN
                assert not balance[col].isna().any(), \
                    f"Colonne {col} de Balance {nom} ne doit pas contenir de NaN"
                
                # Vérifier que toutes les valeurs sont >= 0
                assert (balance[col] >= 0).all(), \
                    f"Colonne {col} de Balance {nom} doit contenir uniquement des valeurs >= 0"
        
        # Vérifier que la colonne Numéro est de type string
        for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
            assert balance['Numéro'].dtype == object or balance['Numéro'].dtype == 'string', \
                f"Colonne Numéro de Balance {nom} doit être de type string, trouvé {balance['Numéro'].dtype}"
            
            # Vérifier qu'il n'y a pas de valeurs vides
            assert not balance['Numéro'].isna().any(), \
                f"Colonne Numéro de Balance {nom} ne doit pas contenir de valeurs vides"
        
        # Vérifier que la colonne Intitulé est de type string
        for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
            assert balance['Intitulé'].dtype == object or balance['Intitulé'].dtype == 'string', \
                f"Colonne Intitulé de Balance {nom} doit être de type string"
        
    finally:
        # Nettoyer le fichier temporaire
        if os.path.exists(fichier_excel):
            try:
                os.unlink(fichier_excel)
            except Exception as e:
                # Ignorer les erreurs de suppression (fichier peut être verrouillé sur Windows)
                pass


def test_property_balance_loading_with_demo_file():
    """
    Test de la propriété avec le fichier de démonstration réel.
    
    **Validates: Requirements 1.1, 1.2**
    
    Ce test vérifie que le fichier de démonstration P000 -BALANCE DEMO N_N-1_N-2.xls
    respecte la propriété de complétude du chargement.
    
    Ce test sert de validation que le fichier de démonstration réel
    respecte les mêmes propriétés que les fichiers générés aléatoirement.
    """
    # Chemin vers le fichier de test
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    fichier_test = os.path.join(base_dir, "P000 -BALANCE DEMO N_N-1_N-2.xls")
    
    # Vérifier que le fichier existe
    if not os.path.exists(fichier_test):
        pytest.skip(f"Fichier de démonstration non trouvé: {fichier_test}")
    
    # Créer le lecteur
    reader = BalanceReader(fichier_test)
    
    # Charger les balances
    balance_n, balance_n1, balance_n2 = reader.charger_balances()
    
    # Vérifier que les 3 balances sont chargées
    assert balance_n is not None
    assert balance_n1 is not None
    assert balance_n2 is not None
    
    # Vérifier que chaque balance a exactement 8 colonnes
    assert len(balance_n.columns) == 8, \
        f"Balance N doit avoir 8 colonnes, trouvé {len(balance_n.columns)}: {list(balance_n.columns)}"
    assert len(balance_n1.columns) == 8, \
        f"Balance N-1 doit avoir 8 colonnes, trouvé {len(balance_n1.columns)}: {list(balance_n1.columns)}"
    assert len(balance_n2.columns) == 8, \
        f"Balance N-2 doit avoir 8 colonnes, trouvé {len(balance_n2.columns)}: {list(balance_n2.columns)}"
    
    # Vérifier les noms de colonnes
    colonnes_attendues = {
        'Numéro', 'Intitulé', 
        'Ant Débit', 'Ant Crédit',
        'Débit', 'Crédit',
        'Solde Débit', 'Solde Crédit'
    }
    
    for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
        colonnes_balance = set(balance.columns)
        assert colonnes_balance == colonnes_attendues, \
            f"Balance {nom}: colonnes attendues {colonnes_attendues}, trouvé {colonnes_balance}"
    
    # Vérifier que tous les montants sont numériques
    colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
    
    for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
        for col in colonnes_montants:
            assert pd.api.types.is_numeric_dtype(balance[col]), \
                f"Colonne {col} de Balance {nom} doit être numérique"
    
    print(f"\n✓ Propriété validée avec le fichier de démonstration")
    print(f"  - Balance N:   {len(balance_n)} comptes, 8 colonnes")
    print(f"  - Balance N-1: {len(balance_n1)} comptes, 8 colonnes")
    print(f"  - Balance N-2: {len(balance_n2)} comptes, 8 colonnes")


@st.composite
def st_column_names_with_spaces(draw):
    """
    Génère des noms de colonnes avec des espaces multiples.
    
    Cette stratégie crée des variations de noms de colonnes avec:
    - Espaces multiples entre les mots
    - Espaces en début et fin
    - Variations aléatoires du nombre d'espaces
    
    Returns:
        List[str]: Liste de noms de colonnes avec espaces multiples
    """
    # Noms de colonnes de base
    base_columns = [
        'Numéro', 'Intitulé', 
        'Ant Débit', 'Ant Crédit',
        'Débit', 'Crédit',
        'Solde Débit', 'Solde Crédit'
    ]
    
    # Générer des variations avec espaces multiples
    columns_with_spaces = []
    for col in base_columns:
        # Nombre d'espaces à ajouter (1 à 5)
        num_spaces = draw(st.integers(min_value=1, max_value=5))
        
        # Remplacer les espaces simples par des espaces multiples
        col_with_spaces = col.replace(' ', ' ' * num_spaces)
        
        # Ajouter des espaces en début/fin aléatoirement
        add_leading = draw(st.booleans())
        add_trailing = draw(st.booleans())
        
        if add_leading:
            leading_spaces = draw(st.integers(min_value=1, max_value=3))
            col_with_spaces = ' ' * leading_spaces + col_with_spaces
        
        if add_trailing:
            trailing_spaces = draw(st.integers(min_value=1, max_value=3))
            col_with_spaces = col_with_spaces + ' ' * trailing_spaces
        
        columns_with_spaces.append(col_with_spaces)
    
    return columns_with_spaces


@st.composite
def st_balance_with_messy_columns(draw):
    """
    Génère un fichier Excel avec des noms de colonnes contenant des espaces multiples.
    
    Returns:
        str: Chemin vers le fichier Excel temporaire créé
    """
    # Générer les noms de colonnes avec espaces multiples
    columns_with_spaces = draw(st_column_names_with_spaces())
    
    # Générer le nombre de comptes
    num_comptes = draw(st.integers(min_value=5, max_value=20))
    
    # Créer un fichier temporaire
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False)
    temp_file.close()
    
    # Créer le workbook
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Créer un seul onglet pour ce test
    ws = wb.create_sheet("BALANCE N")
    
    # En-têtes avec espaces multiples
    ws.append(columns_with_spaces)
    
    # Générer les données
    for i in range(num_comptes):
        # Numéro de compte
        classe = draw(st.sampled_from(['1', '2', '3', '4', '5', '6', '7', '8', '9']))
        sous_classe = draw(st.integers(min_value=0, max_value=9))
        detail = draw(st.integers(min_value=0, max_value=999))
        numero = f"{classe}{sous_classe}{detail:03d}"
        
        # Intitulé
        intitule = f"Compte {numero}"
        
        # Générer les montants
        ant_debit = draw(st.floats(min_value=0, max_value=100000, allow_nan=False, allow_infinity=False))
        ant_credit = draw(st.floats(min_value=0, max_value=100000, allow_nan=False, allow_infinity=False))
        mvt_debit = draw(st.floats(min_value=0, max_value=50000, allow_nan=False, allow_infinity=False))
        mvt_credit = draw(st.floats(min_value=0, max_value=50000, allow_nan=False, allow_infinity=False))
        
        # Calculer les soldes de clôture
        solde_ouverture = ant_debit - ant_credit
        solde_cloture = solde_ouverture + mvt_debit - mvt_credit
        solde_debit = max(0, solde_cloture)
        solde_credit = max(0, -solde_cloture)
        
        # Ajouter la ligne
        ws.append([numero, intitule, ant_debit, ant_credit, 
                  mvt_debit, mvt_credit, solde_debit, solde_credit])
    
    # Sauvegarder le fichier
    wb.save(temp_file.name)
    
    return temp_file.name, columns_with_spaces


# ============================================================================
# PROPERTY-BASED TESTS - COLUMN NORMALIZATION
# ============================================================================

@given(data=st.data())
@settings(max_examples=30, deadline=30000)
def test_property_column_name_normalization(data):
    """
    **Property 2: Column Name Normalization**
    
    **Validates: Requirements 1.4**
    
    For any balance sheet with column names containing multiple spaces, 
    the Balance_Reader must normalize them to single-space format, and 
    this normalization must be idempotent (normalizing twice produces 
    the same result as normalizing once).
    
    This property verifies that:
    1. Column names with multiple spaces are normalized to single spaces
    2. Leading and trailing spaces are removed
    3. Normalization is idempotent (applying twice = applying once)
    4. The normalized column names match the expected standard format
    5. All 8 expected columns are present after normalization
    
    Test Strategy:
    - Generate Excel files with column names containing 1-5 spaces between words
    - Add random leading/trailing spaces to column names
    - Verify that Balance_Reader normalizes all column names correctly
    - Verify that applying normalization twice produces the same result
    """
    # Générer un fichier avec des colonnes mal formatées
    fichier_excel, original_columns = data.draw(st_balance_with_messy_columns())
    
    try:
        # Créer le lecteur
        reader = BalanceReader(fichier_excel)
        
        # Charger le fichier Excel pour obtenir le DataFrame brut
        df_brut = pd.read_excel(fichier_excel, sheet_name="BALANCE N")
        
        # Vérifier que les colonnes originales contiennent bien des espaces multiples
        has_multiple_spaces = any(re.search(r'\s{2,}', str(col)) for col in df_brut.columns)
        has_leading_trailing = any(str(col).strip() != str(col) for col in df_brut.columns)
        
        # Au moins une colonne doit avoir des espaces multiples ou des espaces en début/fin
        assume(has_multiple_spaces or has_leading_trailing)
        
        # Appliquer la normalisation une première fois
        df_normalized_1 = reader.nettoyer_colonnes(df_brut.copy())
        
        # Appliquer la normalisation une deuxième fois (test d'idempotence)
        df_normalized_2 = reader.nettoyer_colonnes(df_normalized_1.copy())
        
        # Vérifier que les colonnes normalisées sont identiques (idempotence)
        assert list(df_normalized_1.columns) == list(df_normalized_2.columns), \
            "La normalisation doit être idempotente (même résultat après 2 applications)"
        
        # Vérifier que toutes les colonnes normalisées n'ont qu'un seul espace entre les mots
        for col in df_normalized_1.columns:
            col_str = str(col)
            # Pas d'espaces multiples
            assert not re.search(r'\s{2,}', col_str), \
                f"La colonne '{col}' contient encore des espaces multiples"
            # Pas d'espaces en début
            assert col_str == col_str.lstrip(), \
                f"La colonne '{col}' contient des espaces en début"
            # Pas d'espaces en fin
            assert col_str == col_str.rstrip(), \
                f"La colonne '{col}' contient des espaces en fin"
        
        # Vérifier que les colonnes attendues sont présentes
        colonnes_attendues = {
            'Numéro', 'Intitulé', 
            'Ant Débit', 'Ant Crédit',
            'Débit', 'Crédit',
            'Solde Débit', 'Solde Crédit'
        }
        
        colonnes_normalisees = set(df_normalized_1.columns)
        assert colonnes_normalisees == colonnes_attendues, \
            f"Colonnes attendues: {colonnes_attendues}, trouvé: {colonnes_normalisees}"
        
        # Vérifier que le nombre de colonnes est correct (8)
        assert len(df_normalized_1.columns) == 8, \
            f"Doit avoir exactement 8 colonnes, trouvé {len(df_normalized_1.columns)}"
        
    finally:
        # Nettoyer le fichier temporaire
        if os.path.exists(fichier_excel):
            try:
                os.unlink(fichier_excel)
            except Exception:
                pass


def test_property_column_normalization_with_demo_file():
    """
    Test de la propriété de normalisation avec le fichier de démonstration.
    
    **Validates: Requirements 1.4**
    
    Ce test vérifie que même si le fichier de démonstration a des colonnes
    bien formatées, la normalisation reste idempotente et ne modifie pas
    les colonnes déjà correctes.
    """
    # Chemin vers le fichier de test
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    fichier_test = os.path.join(base_dir, "P000 -BALANCE DEMO N_N-1_N-2.xls")
    
    # Vérifier que le fichier existe
    if not os.path.exists(fichier_test):
        pytest.skip(f"Fichier de démonstration non trouvé: {fichier_test}")
    
    # Créer le lecteur
    reader = BalanceReader(fichier_test)
    
    # Charger le DataFrame brut
    df_brut = pd.read_excel(fichier_test, sheet_name=0)
    
    # Appliquer la normalisation une première fois
    df_normalized_1 = reader.nettoyer_colonnes(df_brut.copy())
    
    # Appliquer la normalisation une deuxième fois
    df_normalized_2 = reader.nettoyer_colonnes(df_normalized_1.copy())
    
    # Vérifier l'idempotence
    assert list(df_normalized_1.columns) == list(df_normalized_2.columns), \
        "La normalisation doit être idempotente"
    
    # Vérifier qu'aucune colonne n'a d'espaces multiples
    for col in df_normalized_1.columns:
        col_str = str(col)
        assert not re.search(r'\s{2,}', col_str), \
            f"La colonne '{col}' contient des espaces multiples"
        assert col_str == col_str.strip(), \
            f"La colonne '{col}' contient des espaces en début ou fin"
    
    print(f"\n✓ Propriété de normalisation validée avec le fichier de démonstration")
    print(f"  - Colonnes normalisées: {list(df_normalized_1.columns)}")
    print(f"  - Idempotence vérifiée: normalisation(x) = normalisation(normalisation(x))")


if __name__ == "__main__":
    """
    Exécution directe des tests pour validation rapide.
    
    Usage:
        python test_balance_reader_property.py
    """
    print("=" * 70)
    print("PROPERTY-BASED TESTS - BALANCE_READER")
    print("=" * 70)
    
    # Test avec le fichier de démonstration
    print("\n[1] Test avec le fichier de démonstration...")
    try:
        test_property_balance_loading_with_demo_file()
        print("   ✓ Test réussi")
    except Exception as e:
        print(f"   ✗ Test échoué: {str(e)}")
    
    # Test de normalisation avec le fichier de démonstration
    print("\n[2] Test de normalisation avec le fichier de démonstration...")
    try:
        test_property_column_normalization_with_demo_file()
        print("   ✓ Test réussi")
    except Exception as e:
        print(f"   ✗ Test échoué: {str(e)}")
    
    print("\n" + "=" * 70)
    print("Pour exécuter tous les tests property-based avec Hypothesis:")
    print("  pytest test_balance_reader_property.py -v")
    print("=" * 70)
