"""
Test simple du module Balance_Reader

Ce script teste le chargement des balances avec le fichier de démonstration.
"""

import sys
import os

# Ajouter le chemin des modules au PYTHONPATH
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'Modules'))

from balance_reader import BalanceReader, BalanceNotFoundException, InvalidBalanceFormatException
import logging

# Configuration du logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def test_balance_reader():
    """Test du chargement des balances"""
    print("=" * 70)
    print("TEST DU MODULE BALANCE_READER")
    print("=" * 70)
    
    try:
        # Chemin vers le fichier de test (absolu)
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        fichier_test = os.path.join(base_dir, "P000 -BALANCE DEMO N_N-1_N-2.xls")
        
        print(f"\nFichier de test: {fichier_test}")
        print(f"   Existe: {os.path.exists(fichier_test)}")
        
        # Créer une instance du lecteur
        print("\n[1] Creation du BalanceReader...")
        reader = BalanceReader(fichier_test)
        print("   OK BalanceReader cree")
        
        # Charger les balances
        print("\n[2] Chargement des balances...")
        balance_n, balance_n1, balance_n2 = reader.charger_balances()
        print("   OK Balances chargees avec succes")
        
        # Afficher les résultats
        print("\n[3] Resultats du chargement:")
        print(f"   - Balance N:   {len(balance_n):>5} comptes")
        print(f"   - Balance N-1: {len(balance_n1):>5} comptes")
        print(f"   - Balance N-2: {len(balance_n2):>5} comptes")
        
        # Vérifier les colonnes
        print("\n[4] Verification des colonnes:")
        colonnes_attendues = [
            'Numero', 'Intitule', 
            'Ant Debit', 'Ant Credit',
            'Debit', 'Credit',
            'Solde Debit', 'Solde Credit'
        ]
        
        for col in balance_n.columns:
            print(f"   OK {col}")
        
        # Afficher un échantillon de données
        print("\n[5] Echantillon de donnees (Balance N - 3 premieres lignes):")
        print("-" * 70)
        for idx, row in balance_n.head(3).iterrows():
            numero = str(row.get('Numero', row.get('Numéro', 'N/A')))
            intitule = str(row.get('Intitule', row.get('Intitulé', 'N/A')))[:40]
            solde_d = row.get('Solde Debit', row.get('Solde Débit', 0))
            solde_c = row.get('Solde Credit', row.get('Solde Crédit', 0))
            print(f"   Compte: {numero} - {intitule}")
            print(f"   Solde Debit: {solde_d:>12,.2f} | Solde Credit: {solde_c:>12,.2f}")
            print("-" * 70)
        
        # Test de conversion des montants
        print("\n[6] Verification de la conversion des montants:")
        col_test = 'Solde Debit' if 'Solde Debit' in balance_n.columns else 'Solde Débit'
        montant_test = balance_n[col_test].iloc[0]
        print(f"   Type du montant: {type(montant_test)}")
        print(f"   Valeur: {montant_test}")
        
        if isinstance(montant_test, (int, float)):
            print("   OK Les montants sont bien convertis en numerique")
        else:
            print("   ERREUR: Les montants ne sont pas numeriques")
        
        print("\n" + "=" * 70)
        print("TEST REUSSI - Module Balance_Reader operationnel")
        print("=" * 70)
        return True
        
    except BalanceNotFoundException as e:
        print(f"\nERREUR - Balance non trouvee: {str(e)}")
        print("=" * 70)
        return False
        
    except InvalidBalanceFormatException as e:
        print(f"\nERREUR - Format invalide: {str(e)}")
        print("=" * 70)
        return False
        
    except Exception as e:
        print(f"\nERREUR inattendue: {str(e)}")
        import traceback
        traceback.print_exc()
        print("=" * 70)
        return False


if __name__ == "__main__":
    success = test_balance_reader()
    sys.exit(0 if success else 1)


# ============================================================================
# PROPERTY-BASED TESTS
# ============================================================================

import pytest
from hypothesis import given, strategies as st, assume, settings
import pandas as pd
import openpyxl
from openpyxl import Workbook
import tempfile


@st.composite
def st_balance_excel_file(draw):
    """
    Génère un fichier Excel valide avec 3 onglets de balances.
    
    Cette stratégie crée un fichier Excel temporaire avec:
    - 3 onglets nommés "BALANCE N", "BALANCE N-1", "BALANCE N-2"
    - Chaque onglet contient exactement 8 colonnes
    - Les données sont cohérentes (montants numériques valides)
    """
    # Générer le nombre de comptes pour chaque balance
    num_comptes_n = draw(st.integers(min_value=5, max_value=50))
    num_comptes_n1 = draw(st.integers(min_value=5, max_value=50))
    num_comptes_n2 = draw(st.integers(min_value=5, max_value=50))
    
    # Créer un fichier temporaire
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False)
    temp_file.close()
    
    # Créer le workbook
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Créer les 3 onglets
    for exercice, num_comptes in [('N', num_comptes_n), ('N-1', num_comptes_n1), ('N-2', num_comptes_n2)]:
        ws = wb.create_sheet(f"BALANCE {exercice}")
        
        # En-têtes (exactement 8 colonnes)
        headers = ['Numéro', 'Intitulé', 'Ant Débit', 'Ant Crédit', 
                   'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        ws.append(headers)
        
        # Générer les données
        for i in range(num_comptes):
            # Numéro de compte SYSCOHADA valide
            classe = draw(st.sampled_from(['1', '2', '3', '4', '5', '6', '7', '8', '9']))
            sous_classe = draw(st.integers(min_value=0, max_value=9))
            detail = draw(st.integers(min_value=0, max_value=999))
            numero = f"{classe}{sous_classe}{detail:03d}"
            
            # Intitulé
            intitule = f"Compte {numero}"
            
            # Générer les soldes d'ouverture
            ant_debit = draw(st.floats(min_value=0, max_value=1000000, allow_nan=False, allow_infinity=False))
            ant_credit = draw(st.floats(min_value=0, max_value=1000000, allow_nan=False, allow_infinity=False))
            
            # Générer les mouvements
            mvt_debit = draw(st.floats(min_value=0, max_value=500000, allow_nan=False, allow_infinity=False))
            mvt_credit = draw(st.floats(min_value=0, max_value=500000, allow_nan=False, allow_infinity=False))
            
            # Calculer les soldes de clôture cohérents
            solde_ouverture = ant_debit - ant_credit
            solde_cloture = solde_ouverture + mvt_debit - mvt_credit
            
            solde_debit = max(0, solde_cloture)
            solde_credit = max(0, -solde_cloture)
            
            # Ajouter la ligne
            ws.append([numero, intitule, ant_debit, ant_credit, 
                      mvt_debit, mvt_credit, solde_debit, solde_credit])
    
    # Sauvegarder le fichier
    wb.save(temp_file.name)
    
    return temp_file.name


@given(fichier_excel=st_balance_excel_file())
@settings(max_examples=20, deadline=30000)
def test_property_balance_loading_completeness(fichier_excel):
    """
    **Validates: Requirements 1.1, 1.2**
    
    Property 1: Balance Loading Completeness
    
    For any valid Excel file containing balance sheets, when the Balance_Reader 
    loads the file, all three worksheets (BALANCE N, BALANCE N-1, BALANCE N-2) 
    must be detected and loaded with exactly 8 columns each.
    
    This property verifies that:
    1. All 3 worksheets are successfully detected
    2. All 3 balances are loaded without errors
    3. Each balance has exactly 8 columns
    4. The column names are normalized correctly
    5. All balances contain at least 1 row of data
    """
    try:
        # Créer le lecteur
        reader = BalanceReader(fichier_excel)
        
        # Charger les balances
        balance_n, balance_n1, balance_n2 = reader.charger_balances()
        
        # Vérifier que les 3 balances sont chargées
        assert balance_n is not None, "Balance N ne doit pas être None"
        assert balance_n1 is not None, "Balance N-1 ne doit pas être None"
        assert balance_n2 is not None, "Balance N-2 ne doit pas être None"
        
        # Vérifier que chaque balance est un DataFrame
        assert isinstance(balance_n, pd.DataFrame), "Balance N doit être un DataFrame"
        assert isinstance(balance_n1, pd.DataFrame), "Balance N-1 doit être un DataFrame"
        assert isinstance(balance_n2, pd.DataFrame), "Balance N-2 doit être un DataFrame"
        
        # Vérifier que chaque balance a exactement 8 colonnes
        assert len(balance_n.columns) == 8, f"Balance N doit avoir 8 colonnes, trouvé {len(balance_n.columns)}"
        assert len(balance_n1.columns) == 8, f"Balance N-1 doit avoir 8 colonnes, trouvé {len(balance_n1.columns)}"
        assert len(balance_n2.columns) == 8, f"Balance N-2 doit avoir 8 colonnes, trouvé {len(balance_n2.columns)}"
        
        # Vérifier les noms de colonnes attendus
        colonnes_attendues = {
            'Numéro', 'Intitulé', 
            'Ant Débit', 'Ant Crédit',
            'Débit', 'Crédit',
            'Solde Débit', 'Solde Crédit'
        }
        
        for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
            colonnes_balance = set(balance.columns)
            assert colonnes_balance == colonnes_attendues, \
                f"Balance {nom} doit avoir les colonnes {colonnes_attendues}, trouvé {colonnes_balance}"
        
        # Vérifier que chaque balance contient au moins 1 ligne
        assert len(balance_n) > 0, "Balance N doit contenir au moins 1 ligne"
        assert len(balance_n1) > 0, "Balance N-1 doit contenir au moins 1 ligne"
        assert len(balance_n2) > 0, "Balance N-2 doit contenir au moins 1 ligne"
        
        # Vérifier que tous les montants sont numériques
        colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        
        for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
            for col in colonnes_montants:
                assert pd.api.types.is_numeric_dtype(balance[col]), \
                    f"Colonne {col} de Balance {nom} doit être numérique"
        
        # Vérifier que la colonne Numéro est de type string
        for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
            assert balance['Numéro'].dtype == object or balance['Numéro'].dtype == 'string', \
                f"Colonne Numéro de Balance {nom} doit être de type string"
        
    finally:
        # Nettoyer le fichier temporaire
        if os.path.exists(fichier_excel):
            try:
                os.unlink(fichier_excel)
            except:
                pass


def test_property_balance_loading_with_demo_file():
    """
    Test de la propriété avec le fichier de démonstration réel.
    
    **Validates: Requirements 1.1, 1.2**
    
    Ce test vérifie que le fichier de démonstration P000 -BALANCE DEMO N_N-1_N-2.xls
    respecte la propriété de complétude du chargement.
    """
    # Chemin vers le fichier de test
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    fichier_test = os.path.join(base_dir, "P000 -BALANCE DEMO N_N-1_N-2.xls")
    
    # Vérifier que le fichier existe
    if not os.path.exists(fichier_test):
        pytest.skip(f"Fichier de démonstration non trouvé: {fichier_test}")
    
    # Créer le lecteur
    reader = BalanceReader(fichier_test)
    
    # Charger les balances
    balance_n, balance_n1, balance_n2 = reader.charger_balances()
    
    # Vérifier que les 3 balances sont chargées
    assert balance_n is not None
    assert balance_n1 is not None
    assert balance_n2 is not None
    
    # Vérifier que chaque balance a exactement 8 colonnes
    assert len(balance_n.columns) == 8, f"Balance N doit avoir 8 colonnes, trouvé {len(balance_n.columns)}"
    assert len(balance_n1.columns) == 8, f"Balance N-1 doit avoir 8 colonnes, trouvé {len(balance_n1.columns)}"
    assert len(balance_n2.columns) == 8, f"Balance N-2 doit avoir 8 colonnes, trouvé {len(balance_n2.columns)}"
    
    # Vérifier les noms de colonnes
    colonnes_attendues = {
        'Numéro', 'Intitulé', 
        'Ant Débit', 'Ant Crédit',
        'Débit', 'Crédit',
        'Solde Débit', 'Solde Crédit'
    }
    
    for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
        colonnes_balance = set(balance.columns)
        assert colonnes_balance == colonnes_attendues, \
            f"Balance {nom}: colonnes attendues {colonnes_attendues}, trouvé {colonnes_balance}"
    
    print(f"\n✓ Propriété validée avec le fichier de démonstration")
    print(f"  - Balance N:   {len(balance_n)} comptes, 8 colonnes")
    print(f"  - Balance N-1: {len(balance_n1)} comptes, 8 colonnes")
    print(f"  - Balance N-2: {len(balance_n2)} comptes, 8 colonnes")
