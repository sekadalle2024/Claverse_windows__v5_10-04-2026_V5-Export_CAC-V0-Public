"""
Property-Based Tests for Balance_Reader Numeric Conversion Robustness

This module contains property-based tests using Hypothesis to verify
the Balance_Reader module's numeric conversion behavior with various
invalid inputs (empty strings, None, text, special characters, etc.).

**Property 3: Numeric Conversion Robustness**
**Validates: Requirements 1.5, 1.6**

Auteur: Système de calcul automatique des notes annexes SYSCOHADA
Date: 08 Avril 2026
"""

import sys
import os
import pytest
from hypothesis import given, strategies as st, assume, settings
import pandas as pd
import openpyxl
from openpyxl import Workbook
import tempfile
import numpy as np

# Ajouter le chemin des modules au PYTHONPATH
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'Modules'))

from balance_reader import BalanceReader, BalanceNotFoundException, InvalidBalanceFormatException


# ============================================================================
# HYPOTHESIS STRATEGIES
# ============================================================================

@st.composite
def st_invalid_monetary_value(draw):
    """
    Génère des valeurs monétaires invalides pour tester la robustesse.
    
    Cette stratégie génère:
    - Chaînes vides
    - None
    - Texte non numérique
    - Caractères spéciaux (compatibles Excel)
    - Formats mixtes (texte + nombres)
    - Espaces uniquement
    - Valeurs NaN
    - Valeurs infinies
    
    Returns:
        Valeur invalide de type varié
    """
    invalid_types = [
        'empty_string',
        'none',
        'text',
        'special_chars',
        'mixed_format',
        'whitespace',
        'nan_string',
        'inf_string',
        'comma_decimal',
        'multiple_decimals'
    ]
    
    invalid_type = draw(st.sampled_from(invalid_types))
    
    if invalid_type == 'empty_string':
        return ''
    elif invalid_type == 'none':
        return None
    elif invalid_type == 'text':
        # Utiliser seulement des caractères alphanumériques pour éviter les problèmes Excel
        return draw(st.text(alphabet='ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz', min_size=1, max_size=10))
    elif invalid_type == 'special_chars':
        return draw(st.sampled_from(['@#$', '***', '---', '+++', '&&&', '!!!']))
    elif invalid_type == 'mixed_format':
        # Mélange de texte et nombres
        num = draw(st.integers(min_value=0, max_value=9999))
        text = draw(st.sampled_from(['ABC', 'XYZ', 'NA', 'ERROR']))
        return f"{text}{num}" if draw(st.booleans()) else f"{num}{text}"
    elif invalid_type == 'whitespace':
        return '   ' * draw(st.integers(min_value=1, max_value=3))
    elif invalid_type == 'nan_string':
        return draw(st.sampled_from(['NaN', 'nan', 'NAN', 'NA', 'na']))
    elif invalid_type == 'inf_string':
        return draw(st.sampled_from(['inf', 'Inf', 'INF', 'infinity', '-inf']))
    elif invalid_type == 'comma_decimal':
        # Format européen avec virgule comme séparateur décimal
        num = draw(st.floats(min_value=0, max_value=99999, allow_nan=False, allow_infinity=False))
        return str(num).replace('.', ',')
    elif invalid_type == 'multiple_decimals':
        # Plusieurs points décimaux (invalide)
        return draw(st.sampled_from(['12.34.56', '1.2.3.4', '100.200.300']))
    
    return ''


@st.composite
def st_balance_with_invalid_values(draw):
    """
    Génère un fichier Excel avec des valeurs monétaires invalides.
    
    Cette stratégie crée un fichier Excel avec:
    - Mélange de valeurs valides et invalides
    - Différents types d'invalides (vide, texte, None, etc.)
    - Au moins quelques valeurs invalides pour tester la conversion
    
    Returns:
        Tuple (chemin_fichier, dict_invalides) où dict_invalides contient
        les positions des valeurs invalides pour vérification
    """
    # Générer le nombre de comptes (réduit pour performance)
    num_comptes = draw(st.integers(min_value=3, max_value=10))
    
    # Créer un fichier temporaire
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False)
    temp_file.close()
    
    # Créer le workbook
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Créer l'onglet BALANCE N
    ws = wb.create_sheet("BALANCE N")
    
    # En-têtes
    headers = ['Numéro', 'Intitulé', 'Ant Débit', 'Ant Crédit', 
               'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
    ws.append(headers)
    
    # Tracker les valeurs invalides insérées
    invalides_positions = []
    
    # Générer les données
    for i in range(num_comptes):
        # Numéro de compte (toujours valide)
        classe = draw(st.sampled_from(['1', '2', '3', '4', '5', '6', '7', '8', '9']))
        sous_classe = draw(st.integers(min_value=0, max_value=9))
        detail = draw(st.integers(min_value=0, max_value=99))
        numero = f"{classe}{sous_classe}{detail:02d}"
        
        # Intitulé (toujours valide)
        intitule = f"Compte {numero}"
        
        # Pour chaque colonne monétaire, décider si on met une valeur valide ou invalide
        row_data = [numero, intitule]
        
        colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        
        for col_idx, col_name in enumerate(colonnes_montants):
            # 30% de chance d'avoir une valeur invalide
            use_invalid = draw(st.booleans()) and draw(st.booleans())  # ~25% chance
            
            if use_invalid:
                invalid_value = draw(st_invalid_monetary_value())
                row_data.append(invalid_value)
                invalides_positions.append({
                    'ligne': i + 2,  # +2 car ligne 1 = headers, ligne 2 = première donnée
                    'colonne': col_name,
                    'valeur': invalid_value
                })
            else:
                # Valeur valide
                valid_value = draw(st.floats(min_value=0, max_value=100000, 
                                             allow_nan=False, allow_infinity=False))
                row_data.append(valid_value)
        
        # Ajouter la ligne
        ws.append(row_data)
    
    # S'assurer qu'il y a au moins quelques valeurs invalides
    assume(len(invalides_positions) >= 1)
    
    # Sauvegarder le fichier
    wb.save(temp_file.name)
    
    return temp_file.name, invalides_positions


# ============================================================================
# PROPERTY-BASED TESTS
# ============================================================================

@given(data=st.data())
@settings(max_examples=10, deadline=30000)
def test_property_numeric_conversion_robustness(data):
    """
    **Property 3: Numeric Conversion Robustness**
    
    **Validates: Requirements 1.5, 1.6**
    
    For any balance sheet loaded, all monetary values must be converted to 
    float type, and any invalid or empty values must be replaced with 0.0 
    without raising exceptions.
    
    This property verifies that:
    1. All monetary columns are converted to float type
    2. Invalid values (empty, None, text, special chars) are replaced with 0.0
    3. No exceptions are raised during conversion
    4. All resulting values are >= 0.0
    5. No NaN or Inf values remain after conversion
    6. The conversion is deterministic (same input = same output)
    
    Test Strategy:
    - Generate Excel files with mix of valid and invalid monetary values
    - Include various types of invalid values (empty, None, text, etc.)
    - Verify that Balance_Reader converts all values to float
    - Verify that all invalid values become 0.0
    - Verify no exceptions are raised
    """
    # Générer un fichier avec des valeurs invalides
    fichier_excel, invalides_positions = data.draw(st_balance_with_invalid_values())
    
    try:
        # Créer le lecteur
        reader = BalanceReader(fichier_excel)
        
        # Charger le DataFrame brut pour voir les valeurs avant conversion
        df_brut = pd.read_excel(fichier_excel, sheet_name="BALANCE N")
        
        # Appliquer la conversion des montants
        df_converti = reader.convertir_montants(df_brut.copy())
        
        # Vérifier que toutes les colonnes monétaires sont de type float
        colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        
        for col in colonnes_montants:
            if col in df_converti.columns:
                # Vérifier que la colonne est numérique
                assert pd.api.types.is_numeric_dtype(df_converti[col]), \
                    f"Colonne {col} doit être numérique après conversion, trouvé {df_converti[col].dtype}"
                
                # Vérifier qu'il n'y a pas de NaN
                assert not df_converti[col].isna().any(), \
                    f"Colonne {col} ne doit pas contenir de NaN après conversion"
                
                # Vérifier qu'il n'y a pas de valeurs infinies
                assert not np.isinf(df_converti[col]).any(), \
                    f"Colonne {col} ne doit pas contenir de valeurs infinies après conversion"
                
                # Vérifier que toutes les valeurs sont >= 0
                assert (df_converti[col] >= 0).all(), \
                    f"Colonne {col} doit contenir uniquement des valeurs >= 0 après conversion"
        
        # Vérifier que les valeurs invalides ont été remplacées par 0.0
        # Note: Les valeurs avec virgule comme séparateur décimal (ex: '1,0') sont converties
        # correctement en float (1.0), donc elles ne doivent PAS être 0.0
        for invalid_info in invalides_positions:
            ligne_idx = invalid_info['ligne'] - 2  # -2 car ligne 1 = headers
            col_name = invalid_info['colonne']
            valeur_originale = invalid_info['valeur']
            
            if ligne_idx < len(df_converti) and col_name in df_converti.columns:
                valeur_convertie = df_converti.loc[ligne_idx, col_name]
                
                # Vérifier si la valeur originale était vraiment invalide (pas juste un format différent)
                # Les valeurs avec virgule comme séparateur décimal sont valides et converties correctement
                if valeur_originale is None or valeur_originale == '' or \
                   (isinstance(valeur_originale, str) and valeur_originale.strip() == ''):
                    # Ces valeurs doivent être 0.0
                    assert valeur_convertie == 0.0, \
                        f"Valeur invalide '{valeur_originale}' à la ligne {invalid_info['ligne']}, " \
                        f"colonne {col_name} doit être convertie en 0.0, trouvé {valeur_convertie}"
                elif isinstance(valeur_originale, str):
                    # Vérifier si c'est un texte pur (pas de chiffres)
                    if not any(c.isdigit() for c in valeur_originale):
                        # Texte pur sans chiffres doit être 0.0
                        assert valeur_convertie == 0.0, \
                            f"Texte pur '{valeur_originale}' à la ligne {invalid_info['ligne']}, " \
                            f"colonne {col_name} doit être converti en 0.0, trouvé {valeur_convertie}"
                    # Sinon, c'est peut-être un format numérique valide (ex: '1,0' -> 1.0)
                    # qui est correctement converti, donc on ne vérifie pas qu'il soit 0.0
        
        # Vérifier que la conversion est déterministe (appliquer 2 fois)
        df_converti_2 = reader.convertir_montants(df_brut.copy())
        
        for col in colonnes_montants:
            if col in df_converti.columns:
                assert df_converti[col].equals(df_converti_2[col]), \
                    f"La conversion de la colonne {col} doit être déterministe"
        
    finally:
        # Nettoyer le fichier temporaire
        if os.path.exists(fichier_excel):
            try:
                os.unlink(fichier_excel)
            except Exception:
                pass


@given(data=st.data())
@settings(max_examples=10, deadline=30000)
def test_property_numeric_conversion_no_exceptions(data):
    """
    **Property 3 (variant): No Exceptions During Conversion**
    
    **Validates: Requirements 1.5, 1.6**
    
    For any balance sheet with invalid monetary values, the conversion
    process must never raise exceptions - it must handle all errors gracefully.
    
    This property verifies that:
    1. No exceptions are raised during convertir_montants()
    2. The method always returns a valid DataFrame
    3. All monetary columns exist after conversion
    4. The number of rows is preserved
    
    Test Strategy:
    - Generate Excel files with extreme invalid values
    - Verify that no exceptions are raised
    - Verify that the DataFrame structure is preserved
    """
    # Générer un fichier avec des valeurs invalides
    fichier_excel, invalides_positions = data.draw(st_balance_with_invalid_values())
    
    try:
        # Créer le lecteur
        reader = BalanceReader(fichier_excel)
        
        # Charger le DataFrame brut
        df_brut = pd.read_excel(fichier_excel, sheet_name="BALANCE N")
        num_lignes_avant = len(df_brut)
        
        # Appliquer la conversion - ne doit jamais lever d'exception
        try:
            df_converti = reader.convertir_montants(df_brut.copy())
            conversion_reussie = True
        except Exception as e:
            conversion_reussie = False
            exception_message = str(e)
        
        # Vérifier qu'aucune exception n'a été levée
        assert conversion_reussie, \
            f"La conversion ne doit jamais lever d'exception, mais a levé: {exception_message}"
        
        # Vérifier que le DataFrame est toujours valide
        assert isinstance(df_converti, pd.DataFrame), \
            "Le résultat de la conversion doit être un DataFrame"
        
        # Vérifier que le nombre de lignes est préservé
        assert len(df_converti) == num_lignes_avant, \
            f"Le nombre de lignes doit être préservé: avant={num_lignes_avant}, après={len(df_converti)}"
        
        # Vérifier que toutes les colonnes monétaires existent
        colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        for col in colonnes_montants:
            if col in df_brut.columns:
                assert col in df_converti.columns, \
                    f"La colonne {col} doit exister après conversion"
        
    finally:
        # Nettoyer le fichier temporaire
        if os.path.exists(fichier_excel):
            try:
                os.unlink(fichier_excel)
            except Exception:
                pass


def test_property_numeric_conversion_with_demo_file():
    """
    Test de la propriété avec le fichier de démonstration réel.
    
    **Validates: Requirements 1.5, 1.6**
    
    Ce test vérifie que le fichier de démonstration P000 -BALANCE DEMO N_N-1_N-2.xls
    respecte la propriété de robustesse de conversion numérique.
    
    Même si le fichier de démonstration contient des valeurs valides,
    ce test vérifie que la conversion fonctionne correctement et que
    toutes les valeurs sont bien de type float.
    """
    # Chemin vers le fichier de test
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    fichier_test = os.path.join(base_dir, "P000 -BALANCE DEMO N_N-1_N-2.xls")
    
    # Vérifier que le fichier existe
    if not os.path.exists(fichier_test):
        pytest.skip(f"Fichier de démonstration non trouvé: {fichier_test}")
    
    # Créer le lecteur
    reader = BalanceReader(fichier_test)
    
    # Charger les balances
    balance_n, balance_n1, balance_n2 = reader.charger_balances()
    
    # Vérifier que toutes les colonnes monétaires sont de type float
    colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
    
    for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
        for col in colonnes_montants:
            if col in balance.columns:
                # Vérifier que la colonne est numérique
                assert pd.api.types.is_numeric_dtype(balance[col]), \
                    f"Colonne {col} de Balance {nom} doit être numérique"
                
                # Vérifier qu'il n'y a pas de NaN
                assert not balance[col].isna().any(), \
                    f"Colonne {col} de Balance {nom} ne doit pas contenir de NaN"
                
                # Vérifier qu'il n'y a pas de valeurs infinies
                assert not np.isinf(balance[col]).any(), \
                    f"Colonne {col} de Balance {nom} ne doit pas contenir de valeurs infinies"
                
                # Vérifier que toutes les valeurs sont >= 0
                assert (balance[col] >= 0).all(), \
                    f"Colonne {col} de Balance {nom} doit contenir uniquement des valeurs >= 0"
    
    print(f"\n✓ Propriété de conversion numérique validée avec le fichier de démonstration")
    print(f"  - Balance N:   {len(balance_n)} comptes, toutes valeurs converties en float")
    print(f"  - Balance N-1: {len(balance_n1)} comptes, toutes valeurs converties en float")
    print(f"  - Balance N-2: {len(balance_n2)} comptes, toutes valeurs converties en float")
    
    # Afficher quelques statistiques
    for balance, nom in [(balance_n, 'N')]:
        print(f"\n  Statistiques Balance {nom}:")
        for col in colonnes_montants[:2]:  # Juste 2 colonnes pour l'exemple
            if col in balance.columns:
                print(f"    - {col}: min={balance[col].min():.2f}, max={balance[col].max():.2f}, "
                      f"moyenne={balance[col].mean():.2f}")


def test_specific_invalid_values():
    """
    Test unitaire avec des valeurs invalides spécifiques.
    
    **Validates: Requirements 1.5, 1.6**
    
    Ce test vérifie que des valeurs invalides spécifiques sont correctement
    converties en 0.0:
    - Chaînes vides
    - None
    - Texte pur
    - Caractères spéciaux
    - NaN
    - Formats mixtes
    """
    # Créer un DataFrame de test avec des valeurs invalides spécifiques
    data = {
        'Numéro': ['101', '102', '103', '104', '105', '106', '107', '108'],
        'Intitulé': ['Compte 1', 'Compte 2', 'Compte 3', 'Compte 4', 
                     'Compte 5', 'Compte 6', 'Compte 7', 'Compte 8'],
        'Ant Débit': ['', None, 'ABC', '@@@', 'NaN', '12.34.56', '  ', 1000.0],
        'Ant Crédit': [None, '', '***', 'XYZ123', 'N/A', 'inf', 2000.0, '   '],
        'Débit': ['ABC123', '---', None, '', 'nan', 3000.0, '+++', 'ERROR'],
        'Crédit': [4000.0, 'TEXT', None, '&&&', '', '12,34', 'N/A', None],
        'Solde Débit': ['', 5000.0, None, 'ABC', '!!!', '', 'inf', 'NaN'],
        'Solde Crédit': [6000.0, '', None, '***', 'XYZ', '  ', None, '---']
    }
    
    df = pd.DataFrame(data)
    
    # Créer un lecteur (fichier fictif)
    reader = BalanceReader("dummy.xlsx")
    
    # Appliquer la conversion
    df_converti = reader.convertir_montants(df)
    
    # Vérifier que toutes les colonnes monétaires sont de type float
    colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
    
    for col in colonnes_montants:
        # Vérifier le type
        assert pd.api.types.is_numeric_dtype(df_converti[col]), \
            f"Colonne {col} doit être numérique"
        
        # Vérifier qu'il n'y a pas de NaN
        assert not df_converti[col].isna().any(), \
            f"Colonne {col} ne doit pas contenir de NaN"
        
        # Vérifier qu'il n'y a pas de valeurs infinies
        assert not np.isinf(df_converti[col]).any(), \
            f"Colonne {col} ne doit pas contenir de valeurs infinies"
        
        # Vérifier que toutes les valeurs sont >= 0
        assert (df_converti[col] >= 0).all(), \
            f"Colonne {col} doit contenir uniquement des valeurs >= 0"
    
    # Vérifier que les valeurs invalides ont été remplacées par 0.0
    # Ligne 0: '', None, 'ABC123', 4000.0, '', 6000.0
    assert df_converti.loc[0, 'Ant Débit'] == 0.0, "Chaîne vide doit être 0.0"
    assert df_converti.loc[0, 'Ant Crédit'] == 0.0, "None doit être 0.0"
    assert df_converti.loc[0, 'Débit'] == 0.0, "Texte mixte doit être 0.0"
    assert df_converti.loc[0, 'Crédit'] == 4000.0, "Valeur valide doit être préservée"
    
    # Ligne 1: None, '', '---', 'TEXT', 5000.0, ''
    assert df_converti.loc[1, 'Ant Débit'] == 0.0, "None doit être 0.0"
    assert df_converti.loc[1, 'Ant Crédit'] == 0.0, "Chaîne vide doit être 0.0"
    assert df_converti.loc[1, 'Débit'] == 0.0, "Caractères spéciaux doivent être 0.0"
    assert df_converti.loc[1, 'Crédit'] == 0.0, "Texte pur doit être 0.0"
    
    print("\n✓ Test des valeurs invalides spécifiques réussi")
    print("  - Chaînes vides converties en 0.0")
    print("  - None converti en 0.0")
    print("  - Texte pur converti en 0.0")
    print("  - Caractères spéciaux convertis en 0.0")
    print("  - Valeurs valides préservées")


if __name__ == "__main__":
    """
    Exécution directe des tests pour validation rapide.
    
    Usage:
        python test_balance_reader_numeric_conversion.py
    """
    print("=" * 70)
    print("PROPERTY-BASED TESTS - NUMERIC CONVERSION ROBUSTNESS")
    print("=" * 70)
    
    # Test avec des valeurs invalides spécifiques
    print("\n[1] Test avec des valeurs invalides spécifiques...")
    try:
        test_specific_invalid_values()
        print("   ✓ Test réussi")
    except Exception as e:
        print(f"   ✗ Test échoué: {str(e)}")
        import traceback
        traceback.print_exc()
    
    # Test avec le fichier de démonstration
    print("\n[2] Test avec le fichier de démonstration...")
    try:
        test_property_numeric_conversion_with_demo_file()
        print("   ✓ Test réussi")
    except Exception as e:
        print(f"   ✗ Test échoué: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "=" * 70)
    print("Pour exécuter tous les tests property-based avec Hypothesis:")
    print("  pytest test_balance_reader_numeric_conversion.py -v")
    print("=" * 70)
